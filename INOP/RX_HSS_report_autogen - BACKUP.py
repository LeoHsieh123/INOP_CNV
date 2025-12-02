# ========================================================================
# INTEL CONFIDENTIAL
# Copyright 2020 Intel Corporation All Rights Reserved.
#
# The source code contained or described herein and all documents related
# to the source code (Material) are owned by Intel Corporation or its
# suppliers or licensors. Title to the Material remains with Intel Corp-
# oration or its suppliers and licensors. The Material may contain trade
# secrets and proprietary and confidential information of Intel Corpor-
# ation and its suppliers and licensors, and is protected by worldwide
# copyright and trade secret laws and treaty provisions. No part of the
# Material may be used, copied, reproduced, modified, published, uploaded,
# posted, transmitted, distributed, or disclosed in any way without
# Intel's prior express written permission.

# No license under any patent, copyright, trade secret or other intellect-
# ual property right is granted to or conferred upon you by disclosure or
# delivery of the Materials, either expressly, by implication, inducement,
# estoppel or otherwise. Any license under such intellectual property
# rights must be express and approved by Intel in writing.
# ========================================================================
# (1)Identified a bug or need support...
# (2)Improvement request...
# (3)Feedback, comment, or anything else...
# Please email me at yin-jui.chen@intel.com
# ========================================================================
# Please do not forward this to others without the permission of 
# ECS APJ team, thank you
# ========================================================================
# Install Module : type the following command in the terminal
# py -m pip install openpyxl --proxy=http://proxy.png.intel.com:911
# py -m pip install pandas --proxy=http://proxy.png.intel.com:911
# ========================================================================

# Start
print("=============================")
print("RX conformance report\n=============================")
print("Loading module...(1/6)\n=============================")

# Import module
import tkinter as tk
import json, os, shutil
import pandas as pd
import openpyxl as op
from openpyxl.utils import get_column_letter
from shutil import copyfile
from os import walk
from openpyxl.drawing.image import Image
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from tkinter import filedialog
import ANIL_Robot_LHC_Auto_Config as test_cofig
test_speed_selection = test_cofig.speed[:-5]

# Function
def data_append(df,item_1,item_2):
    try:
        df.append(item_1)
    except KeyError:
        df.append(item_2)

def xlsx_general_format(worksheet, start_row, start_column):
    ws = wb[worksheet]
    ws.cell(row=1,column=1).value = "TEST #"
    ws.freeze_panes = "A2"
    
    for i in range(start_row, ws.max_row+1):
        for j in range(start_column, ws.max_column+1):
            if ws.cell(row=i,column=j).value != None:
                ws.cell(row=i,column=j).border = Border(bottom=Side(style='thin'),left=Side(style='thin'),right=Side(style='thin'),top=Side(style='thin'))
            if i == 1:
                ws.cell(row=i,column=j).font = Font(name='Calibri',size=11,bold=True,color='000000')
                ws.cell(row=i,column=j).fill = PatternFill(start_color="D0D0D0", fill_type="solid")
                ws.cell(row=i,column=j).border = Border(bottom=Side(style='double'),left=Side(style='thin'),right=Side(style='thin'),top=Side(style='thin'))
                ws.column_dimensions[get_column_letter(j)].width = 20.62
                ws.column_dimensions[get_column_letter(2)].width = 50.62
            if j == 1 and i !=1:
                ws.cell(row=i,column=j).font = Font(name='Calibri',size=11,bold=True,color='000000')
                ws.cell(row=i,column=j).fill = PatternFill(start_color="AED6F1", fill_type="solid")
            if j == 2 and i !=1:
                ws.cell(row=i,column=j).fill = PatternFill(start_color="FFDCB9", fill_type="solid")
                ws.cell(row=i,column=j).alignment = Alignment(horizontal='left', vertical='center')
            if j != 2:
                ws.cell(row=i,column=j).alignment = Alignment(horizontal='center', vertical='center')

def BER_p_f_format(start_row, start_column):
    ws = wb["BER"]
    for i in range(start_row,ws.max_row+1):
        if ws.cell(row=i,column=start_column).value == "FAIL":
            ws.cell(row=i,column=start_column).font = Font(name='Calibri',size=11,bold=True,color='FFFFFF')
            ws.cell(row=i,column=start_column).fill = PatternFill(start_color="FF0000", fill_type="solid")
        elif ws.cell(row=i,column=start_column).value == "PASS":
            ws.cell(row=i,column=start_column).font = Font(name='Calibri',size=11,bold=True,color='000000')
            ws.cell(row=i,column=start_column).fill = PatternFill(start_color="008000", fill_type="solid")

def BER_error(start_row, start_column):
    ws = wb["BER"]
    for i in range(start_row,ws.max_row+1):
        if type(ws.cell(row=i,column=start_column).value) == (int or float) and ws.cell(row=i,column=start_column).value >1:
            ws.cell(row=i,column=start_column).font = Font(name='Calibri',size=11,bold=True,color='FFFFFF')
            ws.cell(row=i,column=start_column).fill = PatternFill(start_color="FF0000", fill_type="solid")

def BER_CL(start_row, start_column):
    ws = wb["BER"]
    for i in range(start_row,ws.max_row+1):
        #if type(ws.cell(row=i,column=start_column).value) == (int or float) and ws.cell(row=i,column=start_column).value <95:
        if ws.cell(row=i,column=start_column).value <60: #2025/7/14: due to platform performace
            ws.cell(row=i,column=start_column).font = Font(name='Calibri',size=11,bold=True,color='FFFFFF')
            ws.cell(row=i,column=start_column).fill = PatternFill(start_color="FF0000", fill_type="solid")
            ws.cell(row=i,column=start_column-2).font = Font(name='Calibri',size=11,bold=True,color='FFFFFF')
            ws.cell(row=i,column=start_column-2).fill = PatternFill(start_color="FF0000", fill_type="solid")

def BER_speed(start_row, start_column):
    ws = wb["BER"]
    for i in range(start_row,ws.max_row+1):
        if data["DUT"]["INFO"]["CODENAME"] == 'Connorsville' or data["DUT"]["INFO"]["CODENAME"] == 'Granite Rapids-D':
            ws.cell(row=i,column=start_column).font = Font(name='Calibri',size=11,bold=True,color='000000')
            ws.cell(row=i,column=start_column).fill = PatternFill(start_color="FFFFFF", fill_type="solid")
        else:
            if ws.cell(row=i,column=start_column).value == test_speed_selection:
                ws.cell(row=i,column=start_column).font = Font(name='Calibri',size=11,bold=True,color='000000')
                ws.cell(row=i,column=start_column).fill = PatternFill(start_color="FFFFFF", fill_type="solid")
            elif ws.cell(row=i,column=start_column).value != test_speed_selection:
                ws.cell(row=i,column=start_column).font = Font(name='Calibri',size=11,bold=True,color='FFFFFF')
                ws.cell(row=i,column=start_column).fill = PatternFill(start_color="FF0000", fill_type="solid")


def TTL_speed(start_row, start_column, iteration):
    ws = wb["TTL"]
    ws.column_dimensions['I'].width = 16.89
    speed_title=ws.cell(row=start_row,column=start_column).value
    for j in range(iteration):
        ws.cell(row=start_row,column=start_column+j).value = "LINK MODE"
        ws.cell(row=start_row,column=start_column+j).font = Font(name='Calibri',size=11,bold=True,color='000000')
        ws.cell(row=start_row,column=start_column+j).fill = PatternFill(start_color="EBF1DE", fill_type="solid")
        ws.cell(row=start_row,column=start_column+j).border = Border(bottom=Side(style='double'),left=Side(style='thin'),right=Side(style='thin'),top=Side(style='thin'))
        ws.cell(row=start_row,column=start_column+j).alignment = Alignment(horizontal='center', vertical='center')
    for i in range(start_row+1,ws.max_row+1):
        speed_list=ws.cell(row=i,column=start_column).value.split("'"+", "+"'")
        k=len(speed_list)
        for z in range(iteration-k):
            speed_list.append('No Data')
            z=z+1
        for j in range(iteration):
            ws.cell(row=i,column=start_column+j).value=speed_list[j]
            ws.cell(row=i,column=start_column+j).alignment = Alignment(horizontal='center', vertical='center')
            ws.cell(row=i,column=start_column+j).border = Border(bottom=Side(style='thin'),left=Side(style='thin'),right=Side(style='thin'),top=Side(style='thin'))
            if data["DUT"]["INFO"]["CODENAME"] == 'Connorsville' or data["DUT"]["INFO"]["CODENAME"] == 'Granite Rapids-D' :
                ws.cell(row=i,column=start_column+j).font = Font(name='Calibri',size=11,bold=True,color='000000')
                ws.cell(row=i,column=start_column+j).fill = PatternFill(start_color="FFFFFF", fill_type="solid")
            else:
                if ws.cell(row=i,column=start_column+j).value == test_speed_selection:
                    ws.cell(row=i,column=start_column+j).font = Font(name='Calibri',size=11,bold=True,color='000000')
                    ws.cell(row=i,column=start_column+j).fill = PatternFill(start_color="FFFFFF", fill_type="solid")
                elif ws.cell(row=i,column=start_column+j).value != test_speed_selection:
                    ws.cell(row=i,column=start_column+j).font = Font(name='Calibri',size=11,bold=True,color='FFFFFF')
                    ws.cell(row=i,column=start_column+j).fill = PatternFill(start_color="FF0000", fill_type="solid")


    ws.merge_cells(start_row=1, start_column=9, end_row=1, end_column=8+iteration)

# empty List
data = []; empty = "---"
# list DUT/LP-INFO
dut_info = [];lp_info=[]
# list TTL
ttl_max = [];ttl = [];ttl_mean = [];ttl_min = [];ttl_stddev = [];ttl_speed = [];ttl_fec = [];ttl_retry=[];ttl_config = [];
ttl_eye_height_0=[];ttl_eye_height_1=[];ttl_eye_height_2=[];ttl_eye_height_3=[];
rs_fec_corrected=[];rs_fec_uncorrected=[]
ttl_top_eye_height_0 = []; ttl_top_eye_height_1=[];
ttl_middle_eye_height_0 = []; ttl_middle_eye_height_1=[];
ttl_bottom_eye_height_0 = []; ttl_bottom_eye_height_1=[];
# list BER
ber_config = [];ber_errors = [];ber_measured = [];ber_cl = [];ber_p_f = [];ber_results = [];lp_ber_errors=[];lp_ber_measured=[];ber_speed=[]
# list cable length
len_cable = {"_30m":"L01_30m", "_50m":"L02_50m", "_70m":"L03_70m","_90m":"L04_90m", "_100m":"L05_100m","_120m":"L06_120m","_130m":"L07_130m","_140m":"L08_140m","_150m":"L09_150m"}
f = None; f1 = None

# select the folder
copy_dir_path = "copy_raw_data"
if not os.path.isdir(copy_dir_path):
    os.mkdir(copy_dir_path)
else:
    shutil.rmtree(copy_dir_path)
    os.mkdir(copy_dir_path)

print("Select the raw data path...(2/6)\n=============================")

def new_report(test_report):
    lists = os.listdir(test_report)
    lists.sort(key=lambda fn:os.path.getmtime(test_report + "\\" + fn))
    file_new = os.path.join(test_report,lists[-1])
    return file_new

dir = r"C:/BER/"
raw_data_path = new_report(dir)
print("Raw Data Path is ... "+ raw_data_path + "\n=============================")

cable_length_file_path = str(raw_data_path)+ "\\temp_cable_length.json"

if os.path.exists(cable_length_file_path):
    cable_length_file_path  = os.path.join(raw_data_path, "temp_cable_length.json")
    os.remove(cable_length_file_path)
'''

window = tk.Tk()
window.withdraw()
raw_data_path = filedialog.askdirectory()
'''

# json file (TTL)
print("Finding the .json file...(3/6)\n=============================")
count = 0
for root, dirs, files in walk(raw_data_path):
    for f in files:
        if f.endswith('.json'):
            count+=1
            for i in len_cable:
                if i in f:
                    f1 = f.replace(i,len_cable[i])
                    break
                else:
                    f1 = f
            copyfile(root + "/" + f,copy_dir_path + "/" + f1)
title = f


# Append TTL, BER, SNR data
print("Collecting data...(4/6)\n=============================")
info = ["CODENAME", "DEVICE-ID", "DEVICE-NAME", "ETHAGENT-VERSION", "ETRACK-ID", "FIRMWARE-REVISION", "IQV-COMPILED", "NVM-REVISION", "QV-VERSION"]
count = 0
files = os.listdir(copy_dir_path)

for file_num in range(0,len(files)):
    with open(copy_dir_path+"/"+files[file_num]) as f:
        data = json.load(f)

INFO_column = {"INFO":info, "DUT-INFO":dut_info, "LP-INFO":lp_info}
#TTL_column = {"CONFIG":ttl_config, "ITERATION":ttl_iteration, "MAX":ttl_max, "MEAN":ttl_mean, "MIN":ttl_min, "STDDEV":ttl_stddev, "RETRY-ATTEMPTS":ttl_retry, "LINK-MODE":ttl_speed, "FEC-MODE":ttl_fec,"EYE-HEIGHT L0":ttl_eye_height_0,"EYE-HEIGHT L1":ttl_eye_height_1,"EYE-HEIGHT L2":ttl_eye_height_2,"EYE-HEIGHT L3":ttl_eye_height_3}
if data["DUT"]["INFO"]["CODENAME"] == 'Granite Rapids-D':
    if test_cofig.speed == 'HSS_QSFP':
        #TTL_column = {"CONFIG":ttl_config, "TTL":ttl, "RETRY-ATTEMPTS":ttl_retry, "LINK-MODE":ttl_speed, "FEC-MODE":ttl_fec,"EYE-HEIGHT L0":ttl_eye_height_0,"EYE-HEIGHT L1":ttl_eye_height_1,"EYE-HEIGHT L2":ttl_eye_height_2,"EYE-HEIGHT L3":ttl_eye_height_3}
        TTL_column = {"CONFIG":ttl_config, "TTL":ttl, "RETRY-ATTEMPTS":ttl_retry, "LINK-MODE":ttl_speed, "FEC-MODE":ttl_fec,"TOP-EYE-HEIGHT L0":ttl_top_eye_height_0,"TOP-EYE-HEIGHT L1":ttl_top_eye_height_1,"MIDDLE-EYE-HEIGHT L0":ttl_middle_eye_height_0,"MIDDLE-EYE-HEIGHT L1":ttl_middle_eye_height_1,"BOTTOM-EYE-HEIGHT L0":ttl_bottom_eye_height_0,"BOTTOM-EYE-HEIGHT L1":ttl_bottom_eye_height_1}
    elif test_cofig.speed == 'HSS_SFP':
        TTL_column = {"CONFIG":ttl_config, "TTL":ttl, "RETRY-ATTEMPTS":ttl_retry, "LINK-MODE":ttl_speed, "FEC-MODE":ttl_fec,"EYE-HEIGHT L0":ttl_eye_height_0}
elif data["DUT"]["INFO"]["CODENAME"] == 'Connorsville':
    TTL_column = {"CONFIG":ttl_config, "TTL":ttl, "RETRY-ATTEMPTS":ttl_retry, "LINK-MODE":ttl_speed, "FEC-MODE":ttl_fec,"RS-FEC-CORRECTED-WORDS":rs_fec_corrected,"RS-FEC-UNCORRECTED-WORDS":rs_fec_uncorrected}
else:
    TTL_column = {"CONFIG":ttl_config, "TTL":ttl, "RETRY-ATTEMPTS":ttl_retry, "LINK-MODE":ttl_speed, "FEC-MODE":ttl_fec}

BER_column = {"CONFIG":ber_config, "ERRORS":ber_errors, "MEASURED_BER":ber_measured, "CONFIDENCE":ber_cl,"BER_TEST_SPEED":ber_speed,"PASS_FAIL":ber_results,"LP_ERRORS":lp_ber_errors,"LP_MEASURED_BER":lp_ber_measured}


for i in range(1,int(test_cofig.link_iterations)+1):
    ttl_config.append(files[file_num])
    count+=1
    # INFO data
    if count == 1:
        for j in info:
            if j == "DEVICE-ID":
                try:
                    dut_info.append(data["DUT"]["INFO"]["DEVICE-ID"])
                except:
                    dut_info.append(data["DUT"]["INFO"]["DEVICE_ID"])
                try:
                    lp_info.append(data["LP"]["INFO"]["DEVICE-ID"])
                except:
                    lp_info.append(data["LP"]["INFO"]["DEVICE_ID"])
            elif j == "DEVICE-NAME":
                try:
                    dut_info.append(data["DUT"]["INFO"]["DEVICE-NAME"])
                except:
                    dut_info.append(data["DUT"]["INFO"]["BRANDING_STRING"])
                try:
                    lp_info.append(data["LP"]["INFO"]["DEVICE-NAME"])
                except:
                    lp_info.append(data["LP"]["INFO"]["BRANDING_STRING"])

            elif j == "ETRACK-ID":
                try:
                    dut_info.append(data["DUT"]["INFO"]["ETRACK-ID"])
                except:
                    dut_info.append(data["DUT"]["INFO"]["ETRACK_ID"])
                try:
                    lp_info.append(data["LP"]["INFO"]["ETRACK-ID"])
                except:
                    lp_info.append(data["LP"]["INFO"]["ETRACK_ID"])
            else:
                try:
                    dut_info.append(data["DUT"]["INFO"][j])
                except:
                    dut_info.append(empty)
                try:
                    lp_info.append(data["LP"]["INFO"][j])
                except:
                    lp_info.append(empty)
    # TTL data
        if test_cofig.link_iterations !='0':
            for j in TTL_column:
                if j == "RETRY-ATTEMPTS":
                    if test_cofig.reset_side == 'DUT':
                        TTL_column[j].extend(data["DUT"]["TTL"]["RETRY_ATTEMPTS"])
                    else:
                        TTL_column[j].extend(data["LP"]["TTL"]["RETRY_ATTEMPTS"])
                elif j == "LINK-MODE":
                    TTL_column[j].extend(data["DUT"]["LINK"]["LINK-MODE"])
                    
                elif j == "FEC-MODE":
                    TTL_column[j].extend(data["DUT"]["LINK"]["FEC-MODE"])
                    
                elif j == "TOP-EYE-HEIGHT L0":
                    TTL_column[j].extend(data["DUT"]["LANE"]["0"]["TOP-EYE-HEIGHT"])
                    
                elif j == "MIDDLE-EYE-HEIGHT L0":
                    TTL_column[j].extend(data["DUT"]["LANE"]["0"]["MIDDLE-EYE-HEIGHT"])
                    
                elif j == "BOTTOM-EYE-HEIGHT L0":
                    TTL_column[j].extend(data["DUT"]["LANE"]["0"]["BOTTOM-EYE-HEIGHT"])
                    
                elif j == "TOP-EYE-HEIGHT L1":
                    TTL_column[j].extend(data["DUT"]["LANE"]["1"]["TOP-EYE-HEIGHT"])
                    
                elif j == "MIDDLE-EYE-HEIGHT L1":
                    TTL_column[j].extend(data["DUT"]["LANE"]["1"]["MIDDLE-EYE-HEIGHT"])
                    
                elif j == "BOTTOM-EYE-HEIGHT L1":
                    TTL_column[j].extend(data["DUT"]["LANE"]["1"]["BOTTOM-EYE-HEIGHT"])
                    
                elif j == "TTL":
                    TTL_column[j].extend(data["DUT"]["TTL"]["MAC-TTL(ms)"])
                    
                elif j == "RS-FEC-CORRECTED-WORDS":
                    TTL_column[j].extend(data["DUT"]["LINK"]["RS-FEC-CORRECTED-WORDS"])
                    
                elif j == "RS-FEC-UNCORRECTED-WORDS":
                    TTL_column[j].extend(data["DUT"]["LINK"]["RS-FEC-UNCORRECTED-WORDS"])
                    
                elif j == "CONFIG":
                    pass
                else:
                    try:
                        if test_cofig.reset_side == 'DUT':
                            TTL_column[j].append(data["DUT"]["TTL"]["LINK TTL(MS)"][j])
                        else:
                            TTL_column[j].append(data["LP"]["TTL"]["LINK TTL(MS)"][j])
                    except:
                        if test_cofig.reset_side == 'DUT':
                            TTL_column[j].append(data["DUT"]["TTL"]["MAC-TTL(ms)"][j])
                        else:
                            TTL_column[j].append(data["LP"]["TTL"]["MAC-TTL(ms)"][j])

for i in range(1,int(test_cofig.ber_iterations)+1):
    ber_config.append(files[file_num])
    # BER data
    if test_cofig.ber_iterations !='0':
        try:
            ber_p_f = data["DUT"]["BER"]["PASS_FAIL"]
            ber_p_f = [x for x in ber_p_f if x != "N/A"]
        except:
            pass
        for j in BER_column:
            if j == "CONFIDENCE":
                data_append(BER_column[j],data["RESULTS"]["DUT"]["BER"][j]["MIN"], empty)
            elif j == "BER_TEST_SPEED":
                data_append(BER_column[j],data["DUT"]["LINK"]["LINK-SPEED"][0], empty)
            elif j == "PASS_FAIL":
                if len(ber_p_f) == 0:
                    BER_column[j].append(empty)
                elif "FAIL" in ber_p_f:
                    BER_column[j].append("FAIL")
                else:
                    BER_column[j].append("PASS")
            elif j =="CONFIG":
                pass
            elif j=="LP_ERRORS":
                data_append(BER_column[j],data["LP"]["BER"]["ERRORS"][0], empty)
            elif j=="LP_MEASURED_BER":
                data_append(BER_column[j],data["LP"]["BER"]["MEASURED_BER"][0], empty)
            else:
                data_append(BER_column[j],data["RESULTS"]["DUT"]["BER"][j]["MAX"], empty)

shutil.rmtree(copy_dir_path)
empty_avg = {"CONFIG":empty, "ITERATION":empty, "ERRORS":empty, "MEASURED_BER":empty, "CONFIDENCE":empty,"PASS_FAIL":empty}
data_info = pd.DataFrame(INFO_column)
if test_cofig.link_iterations !='0':
    data_ttl = pd.DataFrame(TTL_column)
if test_cofig.ber_iterations !='0':
    data_ber = pd.DataFrame(BER_column)
#data_ttl = data_ttl._append(data_ttl.mean(axis=0),ignore_index=True)
#data_ber = data_ber._append(empty_avg,ignore_index=True)

writer = pd.ExcelWriter(raw_data_path + "/" + raw_data_path.split("/")[-1] + "_RX_report.xlsx", engine='openpyxl') # pylint: disable=abstract-class-instantiated
if test_cofig.link_iterations !='0':
    if test_cofig.ber_iterations !='0':
        frames = {'INFO': data_info, 'TTL': data_ttl, "BER":data_ber}
    else:
        frames = {'INFO': data_info, 'TTL': data_ttl}
else:
    frames = {'INFO': data_info, "BER":data_ber}
for i in frames:
    df = frames[i]
    # 依照實際行數生成新的 index 名稱
    new_index = ["Test_" + str(k+1) for k in range(len(df))]
    df.index = new_index
    frames[i] = df  # 更新回原 dictionary
for sheet, frame in  frames.items():
    if sheet != "INFO":
        frame.to_excel(writer, sheet_name = sheet, na_rep= "---", index=True, startrow=0)
    else:
        frame.to_excel(writer, sheet_name = sheet, na_rep= "---", index=False, startrow=10)
writer.close()

print("Adjusting format...(5/6)\n=============================")
wb = op.load_workbook(raw_data_path + "/" + raw_data_path.split("/")[-1] + "_RX_report.xlsx")
        
ws = wb["INFO"]
img = Image("thumbnail-removebg-preview.png")       
img.width, img.height = (150,150)
img.anchor = ws.cell(row=2, column=1).coordinate
ws.add_image(img)
ws.merge_cells(start_row=1, start_column=1, end_row=10, end_column=3)
ws.cell(row=1,column=1).value = title 
ws.cell(row=1,column=1).font = Font(name='Calibri',size=36,bold=True)
ws.cell(row=1,column=1).fill = PatternFill(start_color="D0D0D0", fill_type="solid")
ws.cell(row=1,column=1).alignment = Alignment(horizontal='center', vertical='center')

for i in range(11,ws.max_row+1):
    for j in range(1,ws.max_column+1):
        if ws.cell(row=i,column=j).value != None:
            ws.cell(row=i,column=j).border = Border(bottom=Side(style='thin'),left=Side(style='thin'),right=Side(style='thin'),top=Side(style='thin'))
        if i == 11:
            ws.cell(row=i,column=j).border = Border(bottom=Side(style='double'),left=Side(style='thin'),right=Side(style='thin'),top=Side(style='thin'))
            ws.cell(row=i,column=j).fill = PatternFill(start_color="D0D0D0", fill_type="solid")
            ws.column_dimensions[get_column_letter(j)].width = 50.62
        if j == 1 and i !=11:
            ws.cell(row=i,column=1).font = Font(name='Calibri',size=11,bold=True,color='000000')
            ws.cell(row=i,column=1).fill = PatternFill(start_color="AED6F1", fill_type="solid")

if test_cofig.link_iterations !='0':
    xlsx_general_format(worksheet="TTL", start_row=1, start_column=1)
if test_cofig.ber_iterations !='0':
    xlsx_general_format(worksheet="BER", start_row=1, start_column=1)

#if test_cofig.link_iterations !='0':
    #TTL_speed(start_row=1, start_column=9, iteration= int(test_cofig.link_iterations))
#TTL_speed(start_row=1, start_column=9, iteration=data["RESULTS"]["ITERATION"]["ITERATION"]["MAX"])
if test_cofig.ber_iterations !='0':
    BER_speed(start_row=2, start_column=6)
    BER_error(start_row=2, start_column=3)
    BER_CL(start_row=2, start_column=5)
    BER_p_f_format(start_row=2, start_column=7)

if test_cofig.link_iterations !='0':
    ws = wb["TTL"]
    ws.column_dimensions['G'].width = 20
    ws.column_dimensions['H'].width = 20
    ws.column_dimensions['I'].width = 20
    ws.column_dimensions['J'].width = 20
    ws.column_dimensions['K'].width = 20
    ws.column_dimensions['L'].width = 20
    for i in range(3,4):
        ws.cell(row=1,column=i).value = ws.cell(row=1,column=i).value + "[ms]"

wb.save(raw_data_path + "/" + raw_data_path.split("/")[-1] + "_RX_report.xlsx")
print("Done...(6/6)\n=============================")
print('Test Repot is in ... '+ raw_data_path)
print("=============================")